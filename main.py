from pathlib import Path
import openpyxl


def extractWord(word):
    """
    :param word: string
    :return: slovo bez interpunkcie
    """
    out = ""
    chars = ["a", "á", "ä", "b", "c", "č", "d", "ď", "e", "é", "f", "g", "h", "i", "í", "j", "k", "l", "ĺ", "ľ", "m",
             "n", \
             "ň", "o", "ó", "ô", "p", "q", "r", "ŕ", "s", "š", "t", "ť", "u", "ú", "v", "w", "x", "y", "ý", "z", "ž", \
             "A", "Á", "Ä", "B", "C", "Č", "D", "Ď", "E", "É", "F", "G", "H", "I", "Í", "J", "K", "L", "Ĺ", "Ľ", "M",
             "N", \
             "Ň", "O", "Ó", "Ô", "P", "Q", "R", "Ŕ", "S", "Š", "T", "Ť", "U", "Ú", "V", "W", "X", "Y", "Ý", "Z", "Ž"]
    for char in word:
        if (char in chars):
            out = out + char
    return out


def findSections(file, nameOfSection):
    """
    Najde sekciu zacinajucu <Topic id="to63" desc="nameOfSection"/>
    :param file: jedna transkribcia s viacerymi opismi
    :param nameOfSection: meno sekcie
    :return: obsah sekcie
    """
    i = 0;
    start = 0;
    contentOfSection = []
    while (True):
        line = extractWord(file[i])

        if (line == nameOfSection and start == 0):
            start = 1
            i += 1
        elif (line == nameOfSection and start == 1):
            break
        elif (start == 1):
            contentOfSection.append(file[i])
            i += 1
        else:
            i += 1

    return contentOfSection


def findIndexes(topics):
    """
    Na zaciatku najde mena sekcie opisov + id ktore su v transkribcii
    :param topics:
    :return: pole v ktorom su meno sekcie + id
    """
    indexes = []
    for line in topics:
        if (len(line) > 0):
            line = line.split(" ")
            idTo = line[1].split("\"")
            desc = line[2].split("\"")

            id = idTo[1]
            big = desc[1]
            if ("big" in big):
                indexes.append([big, id])
    return indexes


# <Section type="report" startTime="22.86" endTime="25.68" topic="to3">
def findSectionByID(file, indexes):
    i = 0;
    start = 0;
    contentOfSection = []

    string = "topic=" + "\"" + indexes[1] + "\""
    while (i < len(file)):
        line = file[i]
        if (string in line):
            start = 1
            i += 1
        elif (('</Section>' in line) and start == 1):
            break
        elif (start == 1):
            contentOfSection.append(file[i])
            i += 1
        else:
            i += 1

    return contentOfSection


def parseRecording(file, directoryNameSec):
    """
    Rozparsuje transkribcie a kazdy opis obrazka vlozi do daneho suboru s opismi a do druheho opisu zapise meno suboru,
    ktore sa pridalo do suboru s opismi obrazka
    :param file: 1 transkribcia v .trs
    :param directoryNameSec:
    :return:
    """
    text = file.read_text()
    text = text.split("\n")

    topics = findSections(text, "Topics")
    indexes = findIndexes(topics)

    for i in range(len(indexes)):
        contentOfSection = findSectionByID(text, indexes[i])

        if ("_" in indexes[i][0]):  # ak je meno sekcie big_big3
            index = indexes[i][0].split("_")
            indexes[i][0] = index[1]
        # print(directoryNameSec+indexes[i][0])
        with open(directoryNameSec + indexes[i][0], 'a') as fileContent:
            for line in contentOfSection:
                fileContent.write(line + "\n")
            fileContent.write(10 * "1" + "\n")
        with open(directoryNameSec + 'fileName' + indexes[i][0], 'a') as fileNames:
            fileNames.write(file.name + "\n")


def initDirectory(directoryName):
    if (Path(directoryName).exists()):
        files = Path(directoryName).glob('*')
        for file in files:
            file.unlink()
    else:
        Path(directoryName).mkdir(parents=True, exist_ok=True)


def extractRecordings(directoryNameSec):
    directoryName = './EwaDat/'
    filesTRS = Path(directoryName).glob('*.trs')

    for file in filesTRS:
        parseRecording(file, directoryNameSec)


def initializeArray(length):
    array = [0 for i in range(length)]
    return array


def numberOfSentences(rec):
    count = 0
    for line in rec:
        for word in line:
            if (word == "." or word == "?" or word == "!"):
                count += 1
    return count


def makeOneDimArray(rec):
    out = []
    for line in rec:
        for word in line:
            out.append(word)
    return out


def count_words(recording):
    """
    :param recording: text 1 nahravky
    :return: pocet vsetkych slov
    """
    text, i = extractPlainTextFromRec(recording, 0)

    text = makeOneDimArray(text)
    wordsN = len(text)
    sentences = numberOfSentences(text)  # pocet viet
    count = wordsN - sentences
    return count


def count_hez(recording):
    """
    <Event desc="hez"
    :param recording:
    :return:
    """

    count = 0;
    recording = recording.split("\n")
    for line in recording:
        if ('<Event desc="hez"' in line):
            count += 1

    return count


def writeToFile(statisticNumbers, fileName, statisticName):
    """
    :param statisticNumbers: zoznam s vypoctami statistik
    :param statisticName: zoznam s menami statistik
    :param i: poradie nahravky
    :param directoryName: priecinok, kde sa vytvori/prepise subor
    :return:
    """
    name = "./statistics/" + "statistic_" + fileName
    with open(name, 'a') as file:
        for j in range(len(statisticNumbers)):
            if (isinstance(statisticNumbers[j], dict) == False):
                line = str(statisticName[j]) + ": " + str(statisticNumbers[j]) + "\n"
                file.write(line)
            else:
                line = str(statisticName[j])+": "
                if not statisticNumbers[j]:
                    file.write(line + "None"+"\n")
                else:
                    file.write(line + "\n")
                    for key, value in statisticNumbers[j].items():
                        line = 19 * " " + key + ": " + str(value) + "\n"
                        file.write(line)

        file.write("\n")
    return


def count_segments(recording):
    count = 0
    recording = recording.split("\n")
    for line in recording:
        if ("<Sync time=" in line):
            count += 1
    if (count != 0):
        return count
    return 0


def averageLengthRecording(rec):
    count = 0
    recLength = recordingLength(rec)
    countSegments = count_segments(rec)
    if (recLength != 0 and countSegments != 0):
        count = recLength / countSegments
    return count


def recordingLength(rec):
    # <Turn speaker="spk1" startTime="286.52" endTime="312.88">

    recording = rec.split("\n")

    if (len(recording) > 0):
        if (len(recording[0]) == 0):

            line = recording[1].split(" ")
        else:
            line = recording[0].split(" ")

        if (len(line) > 3):
            start = line[2]
            end = line[3]

            start = start.split("=")

            end = end.split("=")

            start = start[1][1:-1]

            end = end[1][1:-2]

            start = float(start)
            end = float(end)

            return end - start

    return 0


def countInterjections(rec):
    rec = rec.split("\n")
    rec = extractLines(rec)
    dict = {}

    for line in rec:
        for word in line:
            if (len(word) > 1 and word[0] == "%" and word[1] != "@"):
                if extractWord(word) not in dict:
                    dict[extractWord(word)] = 1
                else:
                    dict[extractWord(word)] += 1
    return dict


def countWordsInSentenceAverage(recording):
    return count_words(recording) / numberOfSentences(recording)


def hezAverageSentence(recording):
    return count_hez(recording) / numberOfSentences(recording)


def countProlongedWords(recording):
    recording = recording.split("\n")
    recording = extractLines(recording)
    count = 0
    for line in recording:
        for word in line:
            if (":" in word):
                count += 1
    return count


def calculateStatistics(recording, name):
    statisticNameFile = Path('statistiky').read_text()
    statisticName = statisticNameFile.split("\n")

    statisticNumbers = initializeArray(len(statisticName) - 1)

    statisticNumbers[0] = count_hez(recording)
    statisticNumbers[1] = count_words(recording)

    # statisticNumbers[2] = countUniqueWords(recording) //TO DO
    statisticNumbers[4] = count_segments(recording)
    statisticNumbers[3] = round(statisticNumbers[1] / statisticNumbers[4], 3)
    statisticNumbers[5] = round(averageLengthRecording(recording), 3)
    statisticNumbers[6] = round(recordingLength(recording), 3)
    statisticNumbers[7] = round(countWordsInSentenceAverage(recording), 3)
    statisticNumbers[8] = countInterjections(recording)
    statisticNumbers[9] = round(hezAverageSentence(recording), 3)
    statisticNumbers[10] = countProlongedWords(recording)

    writeToFile(statisticNumbers, name, statisticName)


def deleteEmpty(line):
    tmp = []
    for word in line:
        if (len(word) > 0):
            tmp.append(word)
    return tmp


def extractLines(rec):
    out = []
    for i in range(len(rec)):
        tmp = []
        if (len(rec[i]) > 0 and "<" != rec[i][0]):
            tmp = rec[i].split(" ")
            tmp = deleteEmpty(tmp)
            out.append(tmp)
        elif (len(rec[i]) > 0 and '<Event desc="ned"' in rec[i]):
            line = out.pop()
            out.append(line[:-1])
    return out


def deleteProlongations(rec):
    out = []
    for line in rec:
        tmp = []
        for word in line:
            if (("." in word) or ("?" in word) or ("!" in word)):

                temp = extractWord(word[:-1])
                tmp.append(temp + word[-1])
            else:
                tmp.append(extractWord(word))
        out.append(tmp)

    return out


def devideDot(rec):
    out = []
    for line in rec:
        tmp = []
        for word in line:
            if (word == "." or word == "?" or word == "!"):
                tmp.append(word)
            elif (("." in word) or ("?" in word) or ("!" in word)):

                tmp.append(word[:-1])
                tmp.append(word[-1])
            else:
                tmp.append(extractWord(word))
        out.append(tmp)
    return out


def removeByChar(rec, char):
    out = []

    for tmp in rec:
        line = []
        for word in tmp:
            if (len(word) > 0 and char != word[0]):
                line.append(word)

            elif (len(word) > 0 and char == word[0]):
                if (("." in word) or \
                        ("?" in word) or \
                        ("!" in word)):
                    temp = extractWord(word[:-1])
                    line.append(temp + word[-1])
                else:
                    temp = extractWord(word)
                    line.append(temp)
        out.append(line)

    return out


def removePercent(rec):
    out = []
    for line in rec:
        tmp = []
        for word in line:

            if (len(word) > 0 and word[0] != "%"):
                tmp.append(word)
        out.append(tmp)

    return out


def devideBySlash(rec):
    out = []
    for i in range(len(rec)):
        line = []
        for word in rec[i]:
            if ("/" in word):
                tmp = word.split("/")
                tmp2 = extractWord(tmp[1])
                line.append(tmp2)

            else:
                line.append(word)
        out.append(line)
    return out


def deleteByString(rec, string):
    out = []
    for line in rec:
        line = []
        for word in line:
            if (string not in word):
                line.append(word)

        out.append(line)
    return out


def deleteInv(rec):
    out = []
    for line in rec:
        line = []
        for word in line:
            if (("[inv]" in word) and word[0] == "["):
                tmp = word.split("]")
                line.append(tmp[1])
            elif (("[inv]" in word) and word[-1] == "]"):
                tmp = word.split("[")
                line.append(tmp[0])
            else:
                line.append(word)
        out.append(line)
    return out

def joinDots(rec):
    out = []
    for i in range(len(rec)):

        for j in range(len(rec[i])):
            if(rec[i][j] == "." or rec[i][j] == "!" or rec[i][j] == "?"):
                if (j == 0 and j > 0):
                    rec[i-1][:-1] = rec[i-1][:-1]+rec[i][j]
                else:
                    rec[i][j-1] = rec[i][j-1] + rec[i][j]

    return rec

def extractPlainTextFromRec(rec, i):
    rec = rec.split("\n")
    rec = extractLines(rec)
    rec = removePercent(rec)
    rec = devideBySlash(rec)
    rec = removeByChar(rec, "(")

    rec = deleteProlongations(rec)
    rec = devideDot(rec)

    #print(rec)
    i += 1
    out = []
    for line in rec:
        if (len(line) > 0):
            out.append(line)
    return out, i


def extractPlainTextFromFile(file):
    text = file.read_text()
    text = text.split(10 * "1")

    i = 0
    content = []
    for rec in text:
        out, i = extractPlainTextFromRec(rec, i)
        content.append(out)
    return content


def extractPlainTextFromFiles():
    directoryName = "./transcribtions/"

    filesToDo = Path(directoryName).glob('big*')
    for file in filesToDo:

        content = extractPlainTextFromFile(file)

        with open(directoryName + "Plain" + file.name, 'w+') as fileContentPlain:

            for rec in content:
                for line in rec:
                    for word in line:

                        fileContentPlain.write(word + " ")
                    fileContentPlain.write("\n")
                fileContentPlain.write("\n\n")


def createStatistics(file):
    transcribtions = file.read_text()
    transcribtions = transcribtions.split(10 * "1")

    for recording in transcribtions:
        if (len(recording) > 1):
            statistics = calculateStatistics(recording, file.name)


def doStatistics():
    filesToDo = Path("./transcribtions/").glob('big*')

    for file in filesToDo:
        createStatistics(file)

def addNameToPlainText(directoryName):
    filesToDo = Path(directoryName).glob('Plainbig*')
    nameFiles = Path(directoryName).glob('fileNamebig*')

    filesToDo = list(filesToDo)
    nameFiles = list(nameFiles)

    for (author, file) in zip(nameFiles, filesToDo):
        #print(author,file)
        file = Path(file)
        content = file.read_text()
        content = content.split("\n\n")

        author = Path(author)
        authors = author.read_text()
        authors = authors.split("\n")

        #print(authors)
        with open(file.name + "PlusAuthor", 'w+') as fileContentPlain:

            for (rec, auth) in zip(content, authors):
                authPlain = auth.split(".")
                fileContentPlain.write(authPlain[0]+"\n")
                fileContentPlain.write(rec+"\n\n")

def findNameInExcel(name):
    """
    V excel tabulke najde riadok zhodny s parametrom name a vezme z neho udaje Code, Age, Diagnosis, MOCA, Edu, edu years,sex
    :param name: kod-meno hovoriaceho
    :return: 7-miestne pole [Code, Age, Diagnosis, MOCA, Edu, edu years,sex]
    """
    output = ["","","","","","",""]
    dataframe = openpyxl.load_workbook("ewa_2021_12_16_10_29_28.xlsx")
    dataframe1 = dataframe.active
    for i in range(0, dataframe1.max_row):
        cell_obj = dataframe1.cell(row=i+1, column=3)
        if(cell_obj.value == name):
            output = [dataframe1.cell(row=i, column=3).value,\
                      dataframe1.cell(row=i, column=6).value, \
                      dataframe1.cell(row=i, column=7).value, \
                      dataframe1.cell(row=i, column=8).value, \
                      dataframe1.cell(row=i, column=9).value,\
                      dataframe1.cell(row=i, column=10).value,\
                      dataframe1.cell(row=i, column=11).value]
            print(output)
            return output






if __name__ == '__main__':
    directoryName = './transcribtions/'
    init = input("Pridal sa novy subor do databazy ? a/n: ")
    if (init == "a"):
        initDirectory(directoryName)
        extractRecordings(directoryName)
    init = input("Prepocitat statistiky ? a/n: ")
    if (init == "a"):
        initDirectory("./statistics/")
        doStatistics()
    extractPlainTextFromFiles()
    addNameToPlainText(directoryName)
    findNameInExcel()
